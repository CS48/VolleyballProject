# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import random, re, json, requests, numpy, os.path
from openpyxl import Workbook, load_workbook
from os import path



first_name_data = ["Emily", "Hannah", "Madison", "Ashley", "Sarah", "Alexis", "Samantha", "Jessica", "Elizabeth",
                   "Taylor", "Lauren", "Alyssa", "Kayla", "Abigail", "Brianna", "Olivia", "Emma", "Megan", "Grace",
                   "Victoria", "Rachel", "Anna", "Sydney", "Destiny", "Morgan", "Jennifer", "Jasmine", "Haley", "Julia",
                   "Kaitlyn", "Nicole", "Amanda", "Katherine", "Natalie", "Hailey", "Alexandra", "Savannah", "Chloe",
                   "Rebecca", "Stephanie", "Maria"]
last_name_data = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Miller", "Davis", "Garcia", "Rodriguez", "Wilson",
                  "Martinez", "Anderson", "Taylor", "Thomas", "Hernandez", "Moore", "Martin", "Jackson", "Thompson",
                  "White", "Lopez", "Lee", "Gonzalez", "Harris", "Clark", "Lewis", "Robinson", "Walker", "Perez",
                  "Hall", "Young", "Allen", "Sanchez", "Wright", "King", "Scott", "Green", "Baker", "Adams", "Nelson",
                  "Hill", "Ramirez", "Campbell", "Mitchell", "Roberts"]

years = ["Fr", "So", "Jr", "Sr"]
positions = ["S", "OH", "OH2", "Opp", "MB", "MB2"]

# this is what creates each player, currently the stats are determined by the position that is passed in as an argument
# when calling the class. At some point I think I want it to be the other way around. As of right now though, the whole
# game engine is dependent on it being this way. The teams being created in the specific order of the positions list
# above is kinda crucial for now. At some point soon, I will be changing this.

def serialize_playerids_json(list):
    # opening the file in w mode should delete the data, so i don't have to worry about manually overwriting.
    file = 'player_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    with open(os.path.join(filepath, file), 'w') as outfile:
        json.dump(list, outfile)

def serialize_gameids_json(list):
    # opening the file in w mode should delete the data, so i don't have to worry about manually overwriting.
    file = 'game_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    with open(os.path.join(filepath, file), 'w') as outfile:
        json.dump(list, outfile)

def deserialize_playerids_json():
    file = 'player_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    if path.exists(os.path.join(filepath, file)):
        with open(os.path.join(filepath, file)) as json_file:
            data = json.load(json_file)
        return list(data)
    else:
        print("No player id txt file.")

def deserialize_gameids_json():
    file = 'game_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    if path.exists(os.path.join(filepath, file)):
        with open(os.path.join(filepath, file)) as json_file:
            data = json.load(json_file)
        return list(data)
    else:
        print("No game id txt file.")

def generate_id(list):
    if list is None:
        id = random.randint(1, 1000)
        return id
    else:
        pass

    while True:
        id = random.randint(1, 1000)
        if id in list:
            continue
        else:
            return id

def append_id(id, list):
    if list is None:
        list = [id]
        return list
    else:
        list.append(id)
        return list

def delete_id(id, list):
    if id in list:
        delete_this = list.index(id)
        del list[delete_this]
        return list
    else:
        print("Couldn't find id")

class Player:
    def __init__(self, arglist):
        # player info
        self.player_id = arglist[0]
        self.firstname = arglist[1]
        self.lastname = arglist[2]
        self.year = arglist[3]
        self.position = arglist[4]
        # stat tracking
        # offense
        self.kills = 0
        self.attack_errors = 0
        self.total_attacks = 0
        self.service_aces = 0
        self.service_errors = 0
        self.total_serves = 0
        self.bhe = 0
        # defense
        self.digs = 0
        self.receiving_errors = 0
        self.blocks = 0
        self.blocking_errors = 0

        #attributes

        self.attacking = arglist[5]
        self.setting = arglist[6]
        self.serving = arglist[7]
        self.receiving = arglist[8]
        self.reaction = arglist[9]
        self.blocking = arglist[10]
        self.volley_iq = arglist[11]





# This creates a team of six players by taking the positions list above and iterating through each item. It calls the
# player class to initialize a player for each of those position, which as I explained above, determines the attributes
# of that player.
def create_team():
    # checks to see if the "Teams" file exists. Creates one if it doesn't
    if path.exists("Teams.xlsx"):
        pass
    else:
        workbook = Workbook()
        workbook.save(filename="Teams.xlsx")

    # makes sure that we are working in the Teams file
    filename = "Teams.xlsx"
    workbook = load_workbook(filename=filename)

    # asks for a name for the new team, and puts some limitations on the input.
    while True:
        team_name = input("Please enter a team name: ")
        if len(team_name) > 10:
            print("Sorry, the limit is 10 characters. Try again.")
            continue
        elif len(team_name) < 1:
            print("You didn't enter anything. Try again")
            continue
        else:
            # we're happy with the value given.
            # we're ready to exit the loop.
            break
    # the team name ends up as the name of the sheet within the Teams file
    team_sheet = workbook.create_sheet(team_name)

    # format sheet
    team_sheet["A1"] = "player_id"
    team_sheet["B1"] = "first_name"
    team_sheet["C1"] = "last_name"
    team_sheet["D1"] = "year"
    team_sheet["E1"] = "position"
    team_sheet["F1"] = "attacking"
    team_sheet["G1"] = "setting"
    team_sheet["H1"] = "serving"
    team_sheet["I1"] = "receiving"
    team_sheet["J1"] = "reaction"
    team_sheet["K1"] = "blocking"
    team_sheet["L1"] = "volley_iq"

    # generates 5 players (one for every position) with random ratings
    for x in range(0, len(positions)):
        row = x+2

        # create_playerids_txt()
        print("hi")
        playerids = deserialize_playerids_json()
        new_id = generate_id(playerids)
        updated_list = append_id(new_id, playerids)
        serialize_playerids_json(updated_list)

        team_sheet["A%d" % row] = new_id
        team_sheet["B%d" % row] = random.choice(first_name_data)
        team_sheet["C%d" % row] = random.choice(last_name_data)
        team_sheet["D%d" % row] = random.choice(years)
        team_sheet["E%d" % row] = positions[x]
        team_sheet["F%d" % row] = int(numpy.random.normal(65, 2, 1))
        team_sheet["G%d" % row] = int(numpy.random.normal(65, 2, 1))
        team_sheet["H%d" % row] = int(numpy.random.normal(65, 1, 1))
        team_sheet["I%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["J%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["K%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["L%d" % row] = int(numpy.random.normal(50, 15, 1))



    # saves the file, very important
    workbook.save(filename=filename)

    # Now we will to create a team records sheet to store the team's wins and losses
    if path.exists("Team_results.xlsx"):
        pass
    else:
        workbook = Workbook()
        workbook.save(filename="Team_results.xlsx")

    # I want to name the file "Team_results" and the sheet will be the name of the team.
    filename = "Team_results.xlsx"
    workbook = load_workbook(filename=filename)
    result_sheet = workbook.create_sheet(team_name)

    result_sheet["A1"] = "game_id"
    result_sheet["B1"] = "opponent"
    result_sheet["C1"] = "result"

    workbook.save(filename=filename)

    # feedback is good design ;)
    print("Team Successfully created\n")

# when we need to make player objects out of excel data, this is how we do it.
def load_team():
    # first we check that the Team file exists
    if path.exists("Teams.xlsx"):
        # then we make sure we are working in it
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)
        # we check that file to make sure that the team (given as an arg) is one of the sheets
        # within the file.
        while True:
            team_name = input("Which team would you like to load?:")
            if team_name in workbook.sheetnames:
                # If it exists, we select that sheet
                sheet = workbook[team_name]
                # this will be a list of player objects
                team = []
                # for each row in that sheet, which represents a player, we make a player object by
                # putting the data from that row into a list and feeding that list into Player() as
                # an arg. Then we put that player in the team list
                for x in range(2, sheet.max_row + 1):
                    for value in sheet.iter_rows(min_row=x, max_row=x, values_only=True):
                        arglist = list(value)
                    player = Player(arglist)
                    team.append(player)
                # return the team, so that we can do something with it.
                return team, team_name
            else:
                print("That team doesn't exist in the spreadsheet, try again")
                continue

def print_team(team):
    for x in team:
        print(x.firstname, x.lastname)
        print(x.position, "\n")

        print("Attacking:", x.attacking)
        print("Setting:", x.setting)
        print("Serving:", x.serving)
        print("Receiving:", x.receiving)
        print("Reaction:", x.reaction)
        print("Blocking:", x.blocking)
        print("Volley_iq:", x.volley_iq)

# This will delete a sheet of data (effectively a team) from the Team file
def delete_team():
    # Check to make sure the Team file exists
    if path.exists("Teams.xlsx"):
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)

        # Show a list of the current sheets (Teams)
        print("Here are the current teams...")
        print(workbook.sheetnames, "\n")


        while True:
            # enter the name of a team to delete it
            team_name = input("Enter team name to delete:")
            if team_name in workbook.sheetnames:

                # this should update the id list to make sure that it is cleared of the deleted players
                sheet = workbook[team_name]
                ids_to_delete = []
                for x in range(2, sheet.max_row + 1):
                    ids_to_delete.append(sheet["A%d" % x].value)
                id_list = deserialize_playerids_json()
                print(ids_to_delete)
                print(id_list)
                for x in ids_to_delete:
                    id_list = delete_id(x, id_list)
                serialize_playerids_json(id_list)

                workbook.remove(workbook[team_name])
                break
            else:
                print("That team doesn't exist, try again")
                continue
        # save after it's done
        workbook.save(filename=filename)

        if path.exists("Team_results.xlsx"):
            filename = "Team_results.xlsx"
            workbook = load_workbook(filename=filename)
            workbook.remove(workbook[team_name])

            workbook.save(filename=filename)

        # feedback is good design
        print("%s has been deleted." % team_name)
    else:
        print("No teams currently.\n")


def show_teams():
    # for checking the sheets (Teams) in the Teams file
    if path.exists("Teams.xlsx"):
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)

        print(workbook.sheetnames, "\n")
    else:
        print("No teams exist currently.\n")

# for making a coin flip based on probability p. Such as an attack that has a 30% probability to work
def flip(p):
    return 'H' if random.random() < p else 'T'

# this function sucks, but it's a working prototype. You pass in the serve rotations for each team. The server on the
# serving team is accurate, because they will be in serve rotation 1 always. The potential receiver, however, is just
# randomly selected from the 1, 2, and 3 serve rotation positions on the other team. This doesn't account for
# serve receive formations or a smart server who targets a specific area of the court or a weak receiver. I'll try to
# implement those things later when I get around to making a better function for this.
def serve(server, receiver):


    # some bull shit math to eventually determine the chances of a service error, ace, or a ball that is touched by the
    # other team and possibly kept in play
    if receiver.receiving >= server.serving:
        diff = 0
    else:
        diff = server.serving - receiver.receiving

    print("%s %s is serving..." % (server.firstname, server.lastname))
    x = (server.serving * .1) * (diff / 1000)

    if x >= .15:
        ace_chance = x
    else:
        ace_chance = .15

    #serve_error_chance = (100 - server.serving) / 4
    serve_error_chance = .05

    # takes the chances calculated and makes a biased coin flip
    if flip(ace_chance) == 'H':
        print("It's in and no one touches it. It's an ace!")
        return 2, receiver
    if flip(serve_error_chance) == 'H':
        print("It's out of bounds. A service error!")
        return 1, receiver

    return 0, server, receiver


def serve_target(server, serve_rotation):
    potential_targets = []
    dict_keys = ["1", "2", "3", "4", "5", "6"]
    rotation = serve_rotation["Rotation"]

    if rotation == 1:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "MB2", "OH2"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 2:
        for x in dict_keys:
            if serve_rotation[x].position in ("MB", "MB2", "OH"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 3:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "OH", "MB"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 4:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "OH", "MB"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 5:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "OH2", "MB"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 6:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "OH2", "MB2"):
                potential_targets.append(serve_rotation[x])
    else:
        print("rotation problem in serve targeting")

    # order target list from lowing receiving to highest receiving
    for x in range(1, len(potential_targets)):
        current_player = potential_targets[x]
        current_value = potential_targets[x].receiving
        current_pos = x

        while current_pos > 0 and potential_targets[current_pos - 1].receiving > current_value:
            potential_targets[current_pos] = potential_targets[current_pos - 1]
            current_pos = current_pos - 1

        potential_targets[current_pos] = current_player

    if 0 <= server.volley_iq < 40:
        target = random.choices(potential_targets, weights=(20, 20, 60), k=1)
    elif 40 <= server.volley_iq < 60:
        target = random.choices(potential_targets, weights=(30, 50, 20), k=1)
    elif 60 <= server.volley_iq < 80:
        target = random.choices(potential_targets, weights=(50, 30, 20), k=1)
    elif 80 <= server.volley_iq <= 100:
        target = random.choices(potential_targets, weights=(60, 20, 20), k=1)
    else:
        print("couldn't determine best attack chance")

    return target


def bump(passer, setter, serve_or_poss):
    error_chance = (100 - passer.receiving) / 300

    if flip(error_chance) == 'H':
        if serve_or_poss:
            print("%s %s can't handle the serve: it's an ace." % (passer.firstname, passer.lastname))
        else:
            print("%s %s can't handle the attack: it's a kill." % (passer.firstname, passer.lastname))
        return 1, setter
    else:
        print("%s %s receives and passes to %s %s." % (passer.firstname, passer.lastname, setter.firstname, setter.lastname))
        return 0, setter

def set(setter, serve_rotation):
    bhe_chance = (100 - setter.setting) / 400

    potential_targets = []
    dict_keys = ["1", "2", "3", "4", "5", "6"]
    rotation = serve_rotation["Rotation"]

    if rotation == 1:
        for x in dict_keys:
            if serve_rotation[x].position in ("Opp", "MB", "OH"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 2:
        for x in dict_keys:
            if serve_rotation[x].position in ("OH2", "Opp", "MB"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 3:
        for x in dict_keys:
            if serve_rotation[x].position in ("MB2", "OH2", "Opp"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 4:
        for x in dict_keys:
            # This one is OH, MB2, OH2 because the setter is now 4, and MB2/OH2 are the main attack
            # options. However, OH is a back row attack option.
            if serve_rotation[x].position in ("OH", "MB2", "OH2"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 5:
        for x in dict_keys:
            # This one is OH, OH2, MB2 because the setter is now 5, and MB2/OH are the main attack
            # options. However, OH2 is a back row attack option.
            if serve_rotation[x].position in ("OH", "OH2", "MB2"):
                potential_targets.append(serve_rotation[x])
    elif rotation == 6:
        for x in dict_keys:
            # This one is MB, OH, Opp because the setter is now 6, and MB/OH are the main attack
            # options. However, Opp is a back row attack option.
            if serve_rotation[x].position in ("MB", "OH", "Opp"):
                potential_targets.append(serve_rotation[x])
    else:
        print("rotation problem in set targeting")

    # randomly select a target for the set. This is a singular player and not a list because we aren't using random.
    # choices which would allow us to bias the selection. I just want to make sure it works first before bringing IQ
    # into it.
    target = random.choice(potential_targets)

    if flip(bhe_chance) == 'H':
        print("%s %s commits a ball handling error." % (setter.firstname, setter.lastname))
        return 1, target
    else:
        print("%s %s sets to %s %s." % (setter.firstname, setter.lastname, target.firstname, target.lastname))
        return 0, target

# What we want to do here is write a function whereby 1 or two blockers can challenge an attacker at the net. With
# the possible outcomes being a successful block, a blocking error, a contact on a ball that remains in play, and
# no contact at all. If it's no contact at all, we want to print nothing. No need to add anything to the console
# readout in that case.
def block(attacker, serve_rotation):
    success_prob = 0
    error_prob = 0
    attack_score = 0
    block_score = 0
    potential_blockers = []
    # Only the players who are front row in the rotation can block
    dict_keys = ["4", "3", "2"]

    for x in dict_keys:
        potential_blockers.append(serve_rotation[x])

    blocker = random.choice(potential_blockers)

    print("%s goes for the block..." % blocker.lastname)
    # calculating the "attack score" which will ultimately be either 0 (bad),(average), or 2 (good)
    if 0 <= attacker.attacking < 20:
        attack_score = attack_score + 0
    elif 20 <= attacker.attacking < 40:
        attack_score = attack_score + 1
    elif 40 <= attacker.attacking < 60:
        attack_score = attack_score + 2
    elif 60 <= attacker.attacking < 80:
        attack_score = attack_score + 3
    elif 80 <= attacker.attacking <= 100:
        attack_score = attack_score + 4
    else:
        print("attack score error: attack stat not between 0 and 100")

    if 0 <= attacker.volley_iq < 33:
        attack_score = attack_score + 0
    elif 33 <= attacker.volley_iq < 66:
        attack_score = attack_score + 1
    elif 66 <= attacker.volley_iq <= 100:
        attack_score = attack_score + 2
    else:
        print("attack score error: iq stat not between 0 and 100")

    if 0 <= attack_score < 3:
        attack_score = 0
    elif 3 <= attack_score < 5:
        attack_score = 1
    elif attack_score >= 5:
        attack_score = 2
    else:
        print("attack_score problem 3")

    # calculating the "block" which will ultimately be either 0 (bad), 1 (average), or 2 (good)
    if 0 <= blocker.blocking < 20:
        block_score = block_score + 0
    elif 20 <= blocker.blocking < 40:
        block_score = block_score + 1
    elif 40 <= blocker.blocking < 60:
        block_score = block_score + 2
    elif 60 <= blocker.blocking < 80:
        block_score = block_score + 3
    elif 80 <= blocker.blocking <= 100:
        block_score = block_score + 4
    else:
        print("block score error: block stat not between 0 and 100")

    if 0 <= blocker.reaction < 33:
        def_score = block_score + 0
    elif 33 <= blocker.reaction < 66:
        def_score = block_score + 1
    elif 66 <= blocker.reaction <= 100:
        block_score = block_score + 2
    else:
        print("block score error: react stat not between 0 and 100")

    if 0 <= block_score < 3:
        block_score = 0
    elif 3 <= block_score < 5:
        block_score = 1
    elif block_score >= 5:
        block_score = 2
    else:
        print("block_score problem 3")

    # here's where the comparison happens and a success_prob is given
    if attack_score == 0 and block_score == 0:
        success_prob = .50
    elif attack_score == 0 and block_score == 1:
        success_prob = .45
    elif attack_score == 0 and block_score == 2:
        success_prob = .40
    elif attack_score == 1 and block_score == 0:
        success_prob = .55
    elif attack_score == 1 and block_score == 1:
        success_prob = .50
    elif attack_score == 1 and block_score == 2:
        success_prob = .45
    elif attack_score == 2 and block_score == 0:
        success_prob = .60
    elif attack_score == 2 and block_score == 1:
        success_prob = .55
    elif attack_score == 2 and block_score == 2:
        success_prob = .50
    else:
        print("bro we got a problem with the block")

    # I decided to create an error probability based on the blocker's block category
    if block_score == 0:
        error_prob = .20
    elif block_score == 1:
        error_prob = .15
    elif block_score == 2:
        error_prob = .10
    else:
        print("bro we got a problem with the block error")

    if flip(error_prob) == 'H':
        print("It's a blocking error")
        return 5, blocker
    elif flip(success_prob) == 'H':
        print("It's blocked successfully!")
        return 4, blocker
    else:
        print("The ball gets past the block attempt.")
        return 3, blocker



# I recently rewrote this function to see how well it would work. Basically, the attacker's attack and iq stats are
# combined and then put into a qualitative category. The same happens with the defender's receive and reaction stats.
# From there. The qualitative categories are compared and the attack is given a success probability. In a future update
# I may play around with having the success probability be a random number in a range rather than a set number for a
# specific category comparison.
def attack(attacker, receiver, serve_rotation):
    attack_score = 0
    def_score = 0
    success_prob = 0
    error_prob = 0

    # things to say before attack to make it look better in the console printout
    phrases = ["%s %s attacks fiercely!", "%s %s attacks!", "%s %s attacks with a big swing!",
               "%s %s manages to get the ball over"]

    # spit out a random attack phrase
    print(random.choice(phrases) % (attacker.firstname, attacker.lastname))

    block_result = block(attacker, serve_rotation)

    if block_result[0] == 5:
        return 4, attacker, block_result[1]
    elif block_result[0] == 4:
        return 3, attacker, block_result[1]
    else:
        # calculating the "attack score" which will ultimately be either 0 (bad),(average), or 2 (good)
        if 0 <= attacker.attacking < 20:
            attack_score = attack_score + 0
        elif 20 <= attacker.attacking < 40:
            attack_score = attack_score + 1
        elif 40 <= attacker.attacking < 60:
            attack_score = attack_score + 2
        elif 60 <= attacker.attacking < 80:
            attack_score = attack_score + 3
        elif 80 <= attacker.attacking <= 100:
            attack_score = attack_score + 4
        else:
            print("attack score error: attack stat not between 0 and 100")

        if 0 <= attacker.volley_iq < 33:
            attack_score = attack_score + 0
        elif 33 <= attacker.volley_iq < 66:
            attack_score = attack_score + 1
        elif 66 <= attacker.volley_iq <= 100:
            attack_score = attack_score + 2
        else:
            print("attack score error: iq stat not between 0 and 100")


        if 0 <= attack_score < 3:
            attack_score = 0
        elif 3 <= attack_score < 5:
            attack_score = 1
        elif attack_score >= 5:
            attack_score = 2
        else:
            print("attack_score problem 3")

        # calculating the "def" which will ultimately be either 0 (bad), 1 (average), or 2 (good)
        if 0 <= receiver.receiving < 20:
            def_score = def_score + 0
        elif 20 <= receiver.receiving < 40:
            def_score = def_score + 1
        elif 40 <= receiver.receiving < 60:
            def_score = def_score + 2
        elif 60 <= receiver.receiving < 80:
            def_score = def_score + 3
        elif 80 <= receiver.receiving <= 100:
            def_score = def_score + 4
        else:
            print("def score error: rec stat not between 0 and 100")

        if 0 <= receiver.reaction < 33:
            def_score = def_score + 0
        elif 33 <= receiver.reaction < 66:
            def_score = def_score + 1
        elif 66 <= receiver.reaction <= 100:
            def_score = def_score + 2
        else:
            print("def score error: react stat not between 0 and 100")

        if 0 <= def_score < 3:
            def_score = 0
        elif 3 <= def_score < 5:
            def_score = 1
        elif def_score >= 5:
            def_score = 2
        else:
            print("def_score problem 3")

        # here's where the comparison happens and a success_prob is given
        if attack_score == 0 and def_score == 0:
            success_prob = .50
        elif attack_score == 0 and def_score == 1:
            success_prob = .45
        elif attack_score == 0 and def_score == 2:
            success_prob = .40
        elif attack_score == 1 and def_score == 0:
            success_prob = .55
        elif attack_score == 1 and def_score == 1:
            success_prob = .50
        elif attack_score == 1 and def_score == 2:
            success_prob = .45
        elif attack_score == 2 and def_score == 0:
            success_prob = .60
        elif attack_score == 2 and def_score == 1:
            success_prob = .55
        elif attack_score == 2 and def_score == 2:
            success_prob = .50
        else:
            print("bro we got a problem with the attack")

        # I decided to create an error probability based on the attacker's attack category
        if attack_score == 0:
            error_prob = .20
        elif attack_score == 1:
            error_prob = .15
        elif attack_score == 2:
            error_prob = .10
        else:
            print("bro we got a problem with the error")

        # make some biased coin flips to determine whether the attack is an error, successful, or is touched by the other
        # team and possibly kept in play
        if flip(error_prob) == 'H':
            print("It's an attacking error")
            return 2, attacker, receiver
        elif flip(success_prob) == 'H':
            print("It's a kill")
            return 1, attacker, receiver
        else:
            return 0, attacker, receiver


# this is what is used to select the players on the other team that should be in defensive receiving positions.
# in total there are three players that could be attacked. I went a step further and used the attacker's volley_iq to
# make the selection of a target weighted. If the attacker has a high knowledge, they have a greater chance of attacker
# the defender with the lowest receiving rating
def attack_target(attacker, serve_rotation):
    potential_targets = []
    dict_keys = ["1", "2", "3", "4", "5", "6"]

    for x in dict_keys:
        if serve_rotation[x].position in ("Opp", "MB2", "OH2"):
            potential_targets.append(serve_rotation[x])

    # order target list from lowing receiving to highest receiving
    for x in range(1, len(potential_targets)):
        current_player = potential_targets[x]
        current_value = potential_targets[x].receiving
        current_pos = x

        while current_pos > 0 and potential_targets[current_pos - 1].receiving > current_value:
            potential_targets[current_pos] = potential_targets[current_pos - 1]
            current_pos = current_pos - 1

        potential_targets[current_pos] = current_player

    if 0 <= attacker.volley_iq < 40:
        target = random.choices(potential_targets, weights=(20, 20, 60), k=1)
    elif 40 <= attacker.volley_iq < 60:
        target = random.choices(potential_targets, weights=(30, 50, 20), k=1)
    elif 60 <= attacker.volley_iq < 80:
        target = random.choices(potential_targets, weights=(50, 30, 20), k=1)
    elif 80 <= attacker.volley_iq <= 100:
        target = random.choices(potential_targets, weights=(60, 20, 20), k=1)
    else:
        print("couldn't determine best attack chance")

    return target





# This is where the magic happens. A beautiful recursive mess of a function that I barely understand.
# Explanation of the arguments:
# hometeam and awayteam = a list of player instances generated by "create_team" function.
# serve_rotation_home and serve_rotation_away = a dictionary that contains the current rotation number along with the
# players in each rotation position. This is very important for knowing the active players and who is serving.
# homescore and awayscore = self explanatory, just pass in zeroes on initial call
# homeserve, awayserve, home_possession, away_possession = only one of these should be true at a time, as they determine
# what branch of the playset function will be executed on any given recursion. If you want the home team to serve first,
# pass in True for homeserve. If you want away team to serve first, pass in True for awayserve. both possession argument
# should be false on initial call.
# home_firstserve, away_firstserve = Due to how rotation works in VB, we need to keep track of whether it is each team's
# first serve so that we don't rotate their serve rotation until after they have lost their first serve. Just pass in
# True for both.
# wasithomeserve = this is just to keep track of whose initial serve it was on a volley. Another piece of info in
# deciding whether to rotate.
# serve_or_poss = variable to keep track of whether the last recursion was a serve or a possession. It's how I am
# currently differentiating receiving errors (and therefore aces) from regular kills. Pass in True, initially
# result = this is an optional argument for keeping track of the player who was the target of an action. Leave it blank
# initially.
def playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, homeserve, awayserve, home_firstserve,
            away_firstserve, newrotation, home_possession, away_possession, wasithomeserve, serve_or_poss, result = None):

    if homescore < 25 and awayscore < 25 or abs(homescore - awayscore) < 2:
        if homeserve:
            if newrotation:
                rotation = serve_rotation_home['Rotation']
                if rotation == 6:
                    rotation = 1
                else:
                    rotation = rotation + 1
                serve_rotation_home['Rotation'] = rotation
                print("Home Rotation %d" % serve_rotation_home['Rotation'])
                rotate_serve(serve_rotation_home)

            if home_firstserve:
                home_firstserve = False
            # Do the serve
            target = serve_target(serve_rotation_home["1"], serve_rotation_away)
            serve_result = serve(serve_rotation_home["1"], target[0])
            # stat track: server gets a serve attempt
            serve_rotation_home["1"].total_serves = serve_rotation_home["1"].total_serves + 1

            # if the serve results in an ace
            if serve_result[0] == 2:
                # stat track: server gets a service ace
                serve_rotation_home["1"].service_aces = serve_rotation_home["1"].service_aces + 1
                # increase the homescore by 1
                homescore = homescore + 1
                print(homescore, awayscore)

                # The serving team won the serve, so they won't rotate
                newrotation = False

                # re-enter the loop with homeserve as true, awayserve false
                return playset(hometeam, awayteam, serve_rotation_home,serve_rotation_away, homescore, awayscore, True,
                        False, home_firstserve, away_firstserve, newrotation, False, False, True, True)

            # if the serve results in a service error
            elif serve_result[0] == 1:
                # stat track: server gets a service error
                serve_rotation_home["1"].service_errors = serve_rotation_home["1"].service_errors + 1
                # increase the awayscore by 1
                awayscore = awayscore +1
                print(homescore, awayscore)

                # re-enter the loop with homeserve false, awayserve true, away_firstserve as false
                if away_firstserve:
                    newrotation = False
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                            True, home_firstserve, True, newrotation, False, False, True, True)
                else:
                    newrotation = True
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                            True, home_firstserve, False, newrotation, False, False, True, True)

            # if the result is 0, then the serve was received and now we must go into an away_possession
            else:
                return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                        False, home_firstserve, away_firstserve, newrotation, False, True, True, True, serve_result)

        elif awayserve:

            if newrotation:
                rotation = serve_rotation_away['Rotation']
                if rotation == 6:
                    rotation = 1
                else:
                    rotation = rotation + 1
                serve_rotation_away['Rotation'] = rotation
                print("Away Rotation %d" % serve_rotation_away['Rotation'])
                rotate_serve(serve_rotation_away)

            if away_firstserve:
                away_firstserve = False
            # Do the serve
            target = serve_target(serve_rotation_away["1"], serve_rotation_home)
            serve_result = serve(serve_rotation_away["1"], target[0])
            # stat track: server gets a serve attempt
            serve_rotation_home["1"].total_serves = serve_rotation_home["1"].total_serves + 1

            # if the serve results in an ace
            if serve_result[0] == 2:
                # stat track: server gets a service ace
                serve_rotation_home["1"].service_aces = serve_rotation_home["1"].service_aces + 1

                # increase the awayscore by 1
                awayscore = awayscore + 1
                print(homescore, awayscore)
                # if it was the away team's first serve, it is no longer their first serve, but they will not rotate

                newrotation = False
                # re-enter the loop with awayserve as true, homeserve false, away_firstserve as false
                return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                        True, home_firstserve, away_firstserve, newrotation, False, False, False, True)

            # if the serve results in a service error
            elif serve_result[0] == 1:
                # stat track: server gets a service error
                serve_rotation_home["1"].service_errors = serve_rotation_home["1"].service_errors + 1

                # increase the homescore by 1
                homescore = homescore + 1
                print(homescore, awayscore)

                # re-enter the loop with awayserve false, homeserve true, home_firstserve as false
                if home_firstserve:
                    newrotation = False
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                            True,
                            False, False, away_firstserve, newrotation, False, False, False, True)
                else:
                    newrotation = True
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                            True,
                            False, False, away_firstserve, newrotation, False, False, False, True)

            # if the result is 0, then the serve was received and now we must go into an home_possession
            else:
                return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                        False, home_firstserve, away_firstserve, newrotation, True, False, False, True,  serve_result)

        elif away_possession:
            # attempt to pass
            pass_result = bump(result[1], awayteam[0], serve_or_poss)

            # if pass result == 1, then there was a receiving error and the home team gets a point and serves again
            if pass_result[0] == 1:

                if serve_or_poss == True:
                    # stat track: Receiver gets an error
                    result[2].receiving_errors = result[2].receiving_errors + 1
                    # stat track: server gets a SA
                    result[1].service_aces = result[1].service_aces + 1
                else:
                    # stat track: result[1] is given a kill, this should be the set_result[1] from the previous iter.
                    result[1].kills = result[1].kills + 1

                # stat track: result
                homescore = homescore + 1
                print(homescore, awayscore)

                # if it was homeserve, don't rotate
                if wasithomeserve:
                    newrotation = False
                    # re-enter the loop with homeserve true, awayserve  false, no new rotation
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, True,
                            False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                # if it was awayserve, rotate
                else:
                    newrotation = True
                    # re-enter the loop with homeserve true, awayserve  false, and new rotation
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, True,
                            False, home_firstserve, away_firstserve, newrotation, False, False,wasithomeserve, False)

            # if == 0, then successful pass
            else:

                # stat track: add a dig to the passer
                result[1].digs = result[1].digs + 1

                # attempt to set
                set_result = set(pass_result[1], serve_rotation_away)

                # if ball handling error, home team serves again after gaining point
                if set_result[0] == 1:
                    # stat track: pass_result[1] (setter) is charged with a bhe
                    pass_result[1].bhe = pass_result[1].bhe + 1

                    homescore = homescore + 1
                    print(homescore, awayscore)

                    # if it was homeserve, don't rotate
                    if wasithomeserve:
                        newrotation = False
                    # re-enter the loop with homeserve true, awayserve  false, no new rotation
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                True,
                                False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                    # if it was awayserve, rotate
                    else:
                        newrotation = True
                        # re-enter the loop with homeserve true, awayserve  false, and new rotation
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                True,
                                False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                # if set is successful...attack
                else:
                    target = attack_target(set_result[1], serve_rotation_home)
                    attack_result = attack(set_result[1], target[0], serve_rotation_home)
                    # stat track: the attacker gets an attack attempt
                    set_result[1].total_attacks = set_result[1].total_attacks + 1

                    # home team blocking error, away team gets point and serves
                    if attack_result[0] == 4:
                        # stat track: attack_result[1] is charged with a blocking error
                        attack_result[2].blocking_errors = attack_result[2].blocking_errors + 1
                        # stat track: attack_result[1] is given a kill
                        attack_result[1].kills = attack_result[1].kills + 1

                        print("%s %s is credited with a kill." % (attack_result[1].firstname, attack_result[1].lastname))

                        awayscore = awayscore + 1
                        print(homescore, awayscore)

                        if wasithomeserve:
                            newrotation = True

                            if away_firstserve:
                                newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, False, True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        else:
                            newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, False, True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    # home team successful block, home team gets point and serves
                    elif attack_result[0] == 3:
                        # stat track: attack_result[1] is given a block
                        attack_result[1].blocks = attack_result[1].blocks + 1

                        homescore = homescore + 1
                        print(homescore, awayscore)

                        if wasithomeserve:
                            newrotation = False
                        # re-enter the loop with homeserve true, awayserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    True,
                                    False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        # if it was awayserve, rotate
                        else:
                            newrotation = True
                            # re-enter the loop with homeserve true, awayserve  false, and new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    True,
                                    False, home_firstserve, away_firstserve, newrotation, False, False,wasithomeserve, False)

                    # attacking error, home team gets point and serves again
                    elif attack_result[0] == 2:
                        # stat track: attacker gets charged with attacking error
                        attack_result[1].attack_errors = attack_result[1].attack_errors + 1

                        homescore = homescore + 1
                        print(homescore, awayscore)

                        # if it was homeserve, don't rotate
                        if wasithomeserve:
                            newrotation = False
                        # re-enter the loop with homeserve true, awayserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    True,
                                    False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        # if it was awayserve, rotate
                        else:
                            newrotation = True
                            # re-enter the loop with homeserve true, awayserve  false, and new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    True,
                                    False, home_firstserve, away_firstserve, newrotation, False, False,wasithomeserve, False)

                    # kill, away team gets point and serves again
                    elif attack_result[0] == 1:
                        # stat track: attacker gets a kill
                        attack_result[1].kills = attack_result[1].kills + 1

                        awayscore = awayscore + 1
                        print(homescore, awayscore)

                        # if it is the away team's serve, they don't rotate. If it is, they do.
                        if wasithomeserve:
                            newrotation = True

                            if away_firstserve:
                                newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, False, True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        else:
                            newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, False, True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    # the ball is received by the home team, now it's time for a home possession
                    else:
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                awayscore, False, False, home_firstserve, away_firstserve, newrotation, True, False, wasithomeserve,False, attack_result)

        elif home_possession:
            # attempt to pass
            pass_result = bump(result[2], hometeam[0], serve_or_poss)

            # if pass result == 1, then there was a receiving error and the away team gets a point and serves again
            if pass_result[0] == 1:

                if serve_or_poss == True:
                    # stat track: Receiver gets an error
                    result[2].receiving_errors = result[2].receiving_errors + 1
                    # stat track: server gets a SA
                    result[1].service_aces = result[1].service_aces + 1
                else:
                    # stat track: result[1] is given a kill, this should be the set_result[1] from the previous iter.
                    result[1].kills = result[1].kills + 1

                awayscore = awayscore + 1
                print(homescore, awayscore)

                # if it was the away team's first serve, it is no longer

                if not wasithomeserve:
                    newrotation = False
                    # re-enter the loop with awayserve true, homeserve  false, no new rotation
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                            True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                else:
                    newrotation = True
                    # re-enter the loop with awayserve true, homeserve  false, no new rotation
                    return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore, False,
                            True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

            # if == 0, then successful pass
            else:
                # stat track: add a dig to the passer
                result[1].digs = result[1].digs + 1

                # attempt to set
                set_result = set(pass_result[1], serve_rotation_home)

                # if ball handling error, away team serves after gaining point
                if set_result[0] == 1:
                    # stat track: pass_result[1] (setter) is charged with a bhe
                    pass_result[1].bhe = pass_result[1].bhe + 1

                    awayscore = awayscore + 1
                    print(homescore, awayscore)

                    # they won't rotate
                    if not wasithomeserve:
                        newrotation = False
                        # re-enter the loop with awayserve true, homeserve  false, no new rotation
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                False,
                                True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    else:
                        newrotation = True
                        # re-enter the loop with awayserve true, homeserve  false, no new rotation
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                False,
                                True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                # if set is successful...attack
                else:
                    target = attack_target(set_result[1], serve_rotation_away)
                    attack_result = attack(set_result[1], target[0], serve_rotation_away)
                    # stat track: the attacker gets an attack attempt
                    set_result[1].total_attacks = set_result[1].total_attacks + 1

                    # away team blocking error, home team gets point and serves
                    if attack_result[0] == 4:
                        # stat track: attack_result[2] is charged with a blocking error
                        attack_result[2].blocking_errors = attack_result[2].blocking_errors + 1
                        # stat track: attack_result[1] is given a kill
                        attack_result[1].kills = attack_result[1].kills + 1

                        print("%s %s is credited with a kill." % (attack_result[1].firstname, attack_result[1].lastname))

                        homescore = homescore + 1
                        print(homescore, awayscore)

                        if wasithomeserve:
                            newrotation = False


                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, True, False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        else:
                            newrotation = True

                            if home_firstserve:
                                newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, True, False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)


                    # successful away team block, away team gets point and serves
                    elif attack_result[0] == 3:
                        # stat track: attack_result[1] is added a block
                        attack_result[1].blocks = attack_result[1].blocks + 1

                        awayscore = awayscore + 1
                        print(homescore, awayscore)

                        if not wasithomeserve:
                            newrotation = False
                            # re-enter the loop with awayserve true, homeserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    False,
                                    True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                        else:
                            newrotation = True
                            # re-enter the loop with awayserve true, homeserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    False,
                                    True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    # attacking error, away team gets point and serves again
                    elif attack_result[0] == 2:
                        # stat track: set_result[1] is charged with an attacking error
                        attack_result[1].attack_errors = attack_result[1].attack_errors + 1

                        awayscore = awayscore + 1
                        print(homescore, awayscore)

                        if not wasithomeserve:
                            newrotation = False
                            # re-enter the loop with awayserve true, homeserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    False,
                                    True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                        else:
                            newrotation = True
                            # re-enter the loop with awayserve true, homeserve  false, no new rotation
                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore, awayscore,
                                    False,
                                    True, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    # kill, home team gets point and serves again
                    elif attack_result[0] == 1:
                        # stat track: set_result[1] is given a kill
                        attack_result[1].kills = attack_result[1].kills + 1

                        homescore = homescore + 1
                        print(homescore, awayscore)

                        # if it is the home team's serve, they don't rotate. If it wasn't, they do.
                        if wasithomeserve:
                            newrotation = False


                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, True, False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)
                        else:
                            newrotation = True

                            if home_firstserve:
                                newrotation = False

                            return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                    awayscore, True, False, home_firstserve, away_firstserve, newrotation, False, False, wasithomeserve, False)

                    # the ball is received by the away team, now it's time for an away possession
                    else:
                        return playset(hometeam, awayteam, serve_rotation_home, serve_rotation_away, homescore,
                                awayscore, False, False, home_firstserve, away_firstserve, newrotation, False, True,
                                wasithomeserve, False, attack_result)

        else:
            print("missed serve and possession trees")

    else:
        if homescore > awayscore:
            print("\nSet Score! \nHome:%d Away:%d \n" % (homescore, awayscore))
            return 1
        else:
            print("\nSet Score! \nHome:%d Away:%d \n" % (homescore, awayscore))
            return 2











def play():
    # make sure there is a teams file before we do anything
    if path.exists("Teams.xlsx"):
        pass
    else:
        print("There are no teams currently.")
        return 0

    # work in the teams file
    filename = "Teams.xlsx"
    workbook = load_workbook(filename=filename)
    # let the user select the two teams that are gonna play
    print("Here are the teams that are available:\n")
    show_teams()

    print("Who is the away team?")
    result = load_team()
    away_team = result[0]
    away_name = result[1]

    print("Who is the home team?")
    result = load_team()
    home_team = result[0]
    home_name = result[1]

    home_set_wins = 0
    away_set_wins = 0
    set_number = 0


    serve_rotation_away = {'Rotation': 1, '1': None, '2': None, '3': None, '4': None, '5': None, '6': None}
    serve_rotation_home = {'Rotation': 1, '1': None, '2': None, '3': None, '4': None, '5': None, '6': None}
    for player in away_team:
        if player.position == "S":
            serve_rotation_away['1'] = player
        elif player.position == "OH":
            serve_rotation_away['2'] = player
        elif player.position == "MB":
            serve_rotation_away['3'] = player
        elif player.position == "Opp":
            serve_rotation_away['4'] = player
        elif player.position == "OH2":
            serve_rotation_away['5'] = player
        elif player.position == "MB2":
            serve_rotation_away['6'] = player
        else:
            print("error assigning player to team1 serve rotation")

    for player in home_team:
        if player.position == "S":
            serve_rotation_home['1'] = player
        elif player.position == "OH":
            serve_rotation_home['2'] = player
        elif player.position == "MB":
            serve_rotation_home['3'] = player
        elif player.position == "Opp":
            serve_rotation_home['4'] = player
        elif player.position == "OH2":
            serve_rotation_home['5'] = player
        elif player.position == "MB2":
            serve_rotation_home['6'] = player
        else:
            print("error assigning player to team2 serve rotation")


    #play the game

    while home_set_wins < 3 and away_set_wins < 3:
        # keep track of what set we are on
        set_number = set_number +1

        # recursive function to play a set and get back a winner of the set
        n = playset(away_team, home_team, serve_rotation_away, serve_rotation_home, 0, 0, True, False,
            True,True, False, False, False, True, True, result=None)


        # update the set score
        if n == 1:
            home_set_wins = home_set_wins + 1
        elif n == 2:
            away_set_wins = away_set_wins + 1
        else:
            print("error in assigning set win")



    print("Game Over")

    if home_set_wins == 3:
        print("%s Wins the Game" % home_name)
    else:
        print("%s Wins the Game" % away_name)

    print(home_set_wins, away_set_wins)

    # log game result in the team_results file
    # first we need to give the game an id
    game_ids = deserialize_gameids_json()
    new_id = generate_id(game_ids)
    updated_list = append_id(new_id, game_ids)
    serialize_gameids_json(updated_list)

    # updating team results sheet for away team
    filename = "Team_results.xlsx"
    workbook = load_workbook(filename=filename)
    awayteam_sheet = workbook[away_name]
    x = awayteam_sheet.max_row + 1
    awayteam_sheet["A%d" % x] = new_id
    awayteam_sheet["B%d" % x] = home_name
    if away_set_wins > home_set_wins:
        awayteam_sheet["C%d" % x] = "W"
    else:
        awayteam_sheet["C%d" % x] = "L"
    workbook.save(filename=filename)

    # updating team results sheet for home team
    hometeam_sheet = workbook[home_name]
    x = hometeam_sheet.max_row + 1
    hometeam_sheet["A%d" % x] = new_id
    hometeam_sheet["B%d" % x] = away_name
    if home_set_wins > away_set_wins:
        hometeam_sheet["C%d" % x] = "W"
    else:
        hometeam_sheet["C%d" % x] = "L"
    workbook.save(filename=filename)

def rotate_serve(serve_rotation_dict):
    one = serve_rotation_dict['1']
    two = serve_rotation_dict['2']
    three = serve_rotation_dict['3']
    four = serve_rotation_dict['4']
    five = serve_rotation_dict['5']
    six = serve_rotation_dict['6']

    serve_rotation_dict['1'] = six
    serve_rotation_dict['2'] = one
    serve_rotation_dict['3'] = two
    serve_rotation_dict['4'] = three
    serve_rotation_dict['5'] = four
    serve_rotation_dict['6'] = five

    return serve_rotation_dict


def main():
    # This is effectively a console based menu that keeps running until you exit it
    # the menu options are selected by entering a number that corresponds with that option.
    # Run main() and give it a try.
    while True:
        print("Menu:\n"
              "1. Generate a Team\n"
              "2. Delete a Team\n"
              "3. See Current Teams\n"
              "4. Load Teams\n"
              "5. Play a game\n"
              "6. Exit")

        selection = input("Input a number:")
        if selection == "1":
            print("Creating a team for you...")
            create_team()
            continue
        elif selection == "2":
            delete_team()
            continue
        elif selection == "3":
            show_teams()
            continue
        elif selection == "4":
            print("Current Teams:")
            show_teams()
            team = load_team()
            print_team(team[0])
            print("\n")
            continue
        elif selection == "5":
            play()
            continue
        elif selection == "6":
            s = input("Are you sure (y/n):")
            if s == "y":
                break
            else:
                continue
        else:
            print("Invalid input, try again.\n")
            continue




def main2():
    create_team()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
