import random
from openpyxl import Workbook, load_workbook
from main import generate_id, deserialize_gameids_json, serialize_gameids_json, append_id
from os import path

team_ids = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
            29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]


class Schedule:
    def __init__(self, team_id, slot_number):
        self.team_id = team_id
        self.done = False
        self.games = {}
        x = 1
        while x <= slot_number:
            self.games[x] = None
            x = x + 1


def create_game(slot, awayteam_id, hometeam_id):
    # On the initialization of this class, give the instance a unique id.
    used_game_ids = deserialize_gameids_json()
    new_game_id = generate_id(used_game_ids)

    if path.exists("Games.xlsx"):
        pass
    else:
        workbook = Workbook()
        sheet = workbook["Sheet"]
        sheet["A1"] = "game_id"
        sheet["B1"] = "slot_no"
        sheet["C1"] = "awayteam_id"
        sheet["D1"] = "hometeam_id"
        workbook.save(filename="Games.xlsx")

        # I want to name the file "Team_results" and the sheet will be the name of the team.
    filename = "Games.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook["Sheet"]
    x = sheet.max_row + 1
    sheet["A%d" % x] = new_game_id
    sheet["B%d" % x] = slot
    sheet["C%d" % x] = awayteam_id
    sheet["D%d" % x] = hometeam_id
    workbook.save(filename=filename)

    new_list = append_id(new_game_id, used_game_ids)
    serialize_gameids_json(new_list)

    return new_game_id


def save_schedule(schedule, slots):
    if path.exists("Schedules.xlsx"):
        pass
    else:
        workbook = Workbook()
        workbook.save(filename="Schedules.xlsx")

        # I want to name the file "Team_results" and the sheet will be the name of the team.
    filename = "Schedules.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    x = sheet.max_row + 1
    for i in range(1, slots):
        sheet.cell(row=x, column=i).value = schedule.games[i]

    workbook.save(filename=filename)


def scheduler(team_list, slots):
    print("entered scheduler")
    # create an empty list in order to store the schedule objects, to be iterated through later
    schedule_list = []
    print("created empty schedule list")

    # iterate through the ids in the team list arguments, create a schedule object for each one, add that object to
    # the schedule list
    for team_id in team_list:
        new_schedule = Schedule(team_id, slots)
        schedule_list.append(new_schedule)
    print("created schedule objects for each team")
    print(schedule_list)
    print("\n")

    # iterate through the schedule objects to fill them with games
    for schedule in schedule_list:
        # each schedule is for a team, so get that team's id
        current_team = schedule.team_id
        print("current team being scheduled: %d" % current_team)

        # I want to track if the last 4 games scheduled were home or away
        last_3_location = []

        # each schedule has 30 games, so i'm going to set this x variable for use in a while loop that loops 30 times.
        # x represents the game being scheduled
        x = 1
        while x <= 30:
            print("Scheduling game %d" % x)

            # if there is already a game in the x slot, skip that slot. if not, proceed to schedule a game in that slot
            if schedule.games[x] is None:
                print("Slot is empty")
                while True:
                    # start by randomly choosing another team's schedule
                    prosp_opp = random.choice(schedule_list)
                    print("%d" % prosp_opp.team_id)

                    # if we selected the schedule we are currently working with, select again
                    if prosp_opp.team_id == current_team:
                        continue
                    else:
                        # check to see if the prospective opponent has a game scheduled in the x slot already,
                        # if not, go ahead and schedule x game for both teams. If so, select another prospective opp.
                        if prosp_opp.games[x] is None:
                            print("Prospective opponent found: %d" % prosp_opp.team_id)
                            if len(last_3_location) == 3:
                                if last_3_location.count("H") == 3:
                                    new_game = create_game(x, current_team, prosp_opp.team_id)
                                    location = "A"
                                elif last_3_location.count("A") == 3:
                                    new_game = create_game(x,prosp_opp.team_id, current_team)
                                    location = "H"
                                elif last_3_location[2] == "H":
                                    new_game = create_game(x,prosp_opp.team_id, current_team)
                                    location = "H"
                                else:
                                    new_game = create_game(x, current_team, prosp_opp.team_id)
                                    location = "A"
                            else:
                                new_game = create_game(x,prosp_opp.team_id, current_team)
                                location = "H"

                            if len(last_3_location) == 3:
                                last_3_location.pop(0)
                                last_3_location.append(location)
                            else:
                                last_3_location.append(location)

                            print(last_3_location)

                            schedule.games[x] = new_game
                            prosp_opp.games[x] = new_game
                            print("Game scheduled! ID: %d\n" % new_game)
                            x = x + 1
                            break
                        else:
                            continue
            else:
                x = x + 1
                continue

        schedule.done = True
        save_schedule(schedule, slots)
        print("schedule saved")


def main():
    scheduler(team_ids, 30)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
